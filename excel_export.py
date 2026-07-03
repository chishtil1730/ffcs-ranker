"""Write the merged Slot/Venue/Faculty + ratings dataset to an Excel workbook."""
from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

YELLOW = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


def _autosize(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 45)


def export_workbook(
    matched_df: pd.DataFrame,
    review_df: pd.DataFrame,
    unmatched_df: pd.DataFrame,
    stats: dict,
    output_path: Path,
) -> None:
    """Write Matched Faculty / Needs Review / Unmatched sheets plus a Stats sheet."""
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        matched_df.to_excel(writer, sheet_name="Matched Faculty", index=False)
        review_df.to_excel(writer, sheet_name="Needs Review", index=False)
        unmatched_df.to_excel(writer, sheet_name="Unmatched", index=False)
        pd.DataFrame(list(stats.items()), columns=["Metric", "Value"]).to_excel(
            writer, sheet_name="Stats", index=False
        )

        wb = writer.book
        matched_ws = wb["Matched Faculty"]
        if "Confidence" in matched_df.columns:
            conf_col = matched_df.columns.get_loc("Confidence") + 1
            for row_idx, conf in enumerate(matched_df["Confidence"], start=2):
                if conf is not None and conf < 95:
                    matched_ws.cell(row=row_idx, column=conf_col).fill = YELLOW

        review_ws = wb["Needs Review"]
        for row in review_ws.iter_rows(min_row=2, max_row=review_ws.max_row):
            for cell in row:
                cell.fill = YELLOW

        for sheet in (matched_ws, review_ws, wb["Unmatched"], wb["Stats"]):
            _autosize(sheet)

    logger.info("Workbook written to %s", output_path)
